"""
Microsoft Teams & Outlook MCP Server
MCP server for managing Teams chats/channels and Outlook mail via Microsoft Graph API
"""

import sys
import os
import re
import json
import io
import subprocess
from collections import Counter
from datetime import datetime, timezone, timedelta
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")
import msal
import requests
from importlib.metadata import version as _pkg_version
from fastmcp import FastMCP

__version__ = _pkg_version("ms-teams-mcp")

# ─────────────────────────────────────────
# Configuration (lazy init — prevents crash on import)
# ─────────────────────────────────────────
SCOPES = [
    "Mail.Read", "Mail.Send", "User.Read",
    "Chat.Read", "Chat.ReadWrite",
    "Channel.ReadBasic.All",
    "ChannelMessage.Read.All", "ChannelMessage.Send",
    "Team.ReadBasic.All",
    "Files.Read.All",
    "Sites.Read.All",
    "People.Read",
    "Calendars.ReadWrite",
]
TOKEN_CACHE_FILE = os.path.expanduser(os.environ.get("MS_TOKEN_CACHE", "~/.ms_mcp_token.json"))
FILE_INDEX_PATH = os.path.expanduser("~/.ms_mcp_file_index.json")
GITHUB_REPO = "giljoonseok/ms-teams-mcp"
UPDATE_CHECK_CACHE = os.path.expanduser("~/.ms_mcp_update_check.json")

_cache = None
_app = None
_pub_app = None

def _get_config():
    client_id = os.environ.get("MS_CLIENT_ID")
    client_secret = os.environ.get("MS_CLIENT_SECRET")
    tenant_id = os.environ.get("MS_TENANT_ID")
    if not client_id or not client_secret or not tenant_id:
        raise RuntimeError(
            "Environment variables not set: MS_CLIENT_ID, MS_CLIENT_SECRET, MS_TENANT_ID\n"
            "Check env in MCP client settings, or set environment variables for CLI usage."
        )
    return client_id, client_secret, tenant_id

def _get_cache():
    global _cache
    if _cache is None:
        _cache = msal.SerializableTokenCache()
        try:
            with open(TOKEN_CACHE_FILE, "r") as f:
                _cache.deserialize(f.read())
        except FileNotFoundError:
            pass
    return _cache

def _get_app():
    global _app
    if _app is None:
        client_id, client_secret, tenant_id = _get_config()
        _app = msal.ConfidentialClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            client_credential=client_secret,
            token_cache=_get_cache(),
        )
    return _app

def _get_pub_app():
    global _pub_app
    if _pub_app is None:
        client_id, _, tenant_id = _get_config()
        _pub_app = msal.PublicClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            token_cache=_get_cache(),
        )
    return _pub_app

def save_cache():
    cache = _get_cache()
    if cache.has_state_changed:
        with open(TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())

# ─────────────────────────────────────────
# Token acquisition (silent refresh → error on failure)
# ─────────────────────────────────────────
def _reload_cache():
    """Reload token cache from disk to pick up CLI auth results"""
    cache = _get_cache()
    try:
        with open(TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())
    except FileNotFoundError:
        pass

def get_token():
    # Try ConfidentialClientApplication first, then fall back to PublicClientApplication.
    # CLI auth uses PublicClient, so its cached accounts may not be visible to ConfidentialClient.
    for app_getter in (_get_app, _get_pub_app):
        app = app_getter()
        for attempt in range(2):
            accounts = app.get_accounts()
            if accounts:
                result = app.acquire_token_silent(SCOPES, account=accounts[0])
                if result and "access_token" in result:
                    save_cache()
                    return result["access_token"]
            if attempt == 0:
                _reload_cache()
    raise Exception(
        "Authentication required. Run 'ms-teams-mcp auth' or "
        "call the authenticate tool in MCP."
    )

def _headers():
    token = get_token()
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def _check_response(res: requests.Response):
    """Check Graph API response status and return user-friendly errors"""
    if res.ok:
        return
    status = res.status_code
    try:
        error_data = res.json()
        error_msg = error_data.get("error", {}).get("message", res.text[:200])
    except Exception:
        error_msg = res.text[:200]

    error_map = {
        401: f"Auth error (401): Token expired or invalid. Re-authentication required.\nDetails: {error_msg}",
        403: f"Permission denied (403): Insufficient permissions for this operation. Check Azure AD app permissions.\nDetails: {error_msg}",
        404: f"Resource not found (404): The requested item does not exist or is inaccessible.\nDetails: {error_msg}",
        429: f"Rate limit exceeded (429): Please try again later.\nDetails: {error_msg}",
    }

    raise Exception(error_map.get(status, f"Graph API error ({status}): {error_msg}"))

def graph_get(path: str, params: dict = None, url: str = None):
    if url is None:
        url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.get(url, headers=_headers(), params=params)
    _check_response(res)
    return res.json()

def graph_post(path: str, body: dict):
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.post(url, headers=_headers(), json=body)
    _check_response(res)
    return res.json()

def graph_patch(path: str, body: dict):
    """Graph API PATCH for updating resources"""
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.patch(url, headers=_headers(), json=body)
    _check_response(res)
    return res.json()

def graph_delete(path: str):
    """Graph API DELETE for removing resources"""
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.delete(url, headers=_headers())
    _check_response(res)

def graph_post_action(path: str, body: dict):
    """Graph API POST for 202 No Content responses — sendMail, reply, forward, etc."""
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.post(url, headers=_headers(), json=body)
    _check_response(res)

def _parse_recipients(addresses: str) -> list[dict]:
    """Convert comma-separated email addresses to Graph API recipients format"""
    return [{"emailAddress": {"address": a.strip()}} for a in addresses.split(",") if a.strip()]

def _parse_attendees(addresses: str) -> list[dict]:
    """Convert comma-separated email addresses to Graph API attendees format"""
    return [
        {"emailAddress": {"address": a.strip()}, "type": "required"}
        for a in addresses.split(",") if a.strip()
    ]

def _pagination_footer(data: dict, skip: int, top: int) -> str:
    """Return next-page guidance when more data is available"""
    next_link = data.get("@odata.nextLink", "")
    if next_link:
        return (
            f"\n\n--- More data available ---"
            f"\nNext page: skip={skip + top} or use next_link value"
            f"\nnext_link: {next_link}"
        )
    return ""

_RE_HTML_TAG = re.compile(r"<[^>]+>")
_RE_MULTI_NEWLINE = re.compile(r"\n{3,}")

def strip_html(html: str) -> str:
    if not html:
        return ""
    text = _RE_HTML_TAG.sub("", html).strip()
    return _RE_MULTI_NEWLINE.sub("\n\n", text)

# ─────────────────────────────────────────
# Update Check
# ─────────────────────────────────────────

def _parse_version(version_str: str) -> tuple:
    """Parse version string like 'v0.2.0' or '0.2.0' into comparable tuple"""
    v = version_str.lstrip("v").strip()
    try:
        return tuple(int(x) for x in v.split("."))
    except (ValueError, AttributeError):
        return (0,)

def _check_and_auto_update() -> str | None:
    """Check PyPI for newer version and auto-update if available. Returns status message."""
    try:
        # Check cache — skip if checked within 24 hours
        if os.path.exists(UPDATE_CHECK_CACHE):
            with open(UPDATE_CHECK_CACHE, "r") as f:
                cache = json.load(f)
            last_check = cache.get("last_check", "")
            if last_check:
                last_dt = datetime.fromisoformat(last_check)
                now = datetime.now(timezone.utc)
                if (now - last_dt).total_seconds() < 86400:
                    return cache.get("message")

        # Fetch latest version from PyPI
        resp = requests.get(
            "https://pypi.org/pypi/ms-teams-mcp/json",
            timeout=5,
        )
        if not resp.ok:
            return None

        latest_tag = resp.json().get("info", {}).get("version", "")
        if not latest_tag:
            return None

        latest_ver = _parse_version(latest_tag)
        current_ver = _parse_version(__version__)

        message = None
        if latest_ver > current_ver:
            message = (
                f"[ms-teams-mcp] Updating: {__version__} -> {latest_tag} ..."
            )
            if sys.stderr is not None:
                print(message, file=sys.stderr, flush=True)

            # Auto-update via pip
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--upgrade", "ms-teams-mcp"],
                capture_output=True, text=True, timeout=120,
            )
            if result.returncode == 0:
                message = (
                    f"[ms-teams-mcp] Updated successfully: {__version__} -> {latest_tag}\n"
                    f"  Please restart the server to use the new version."
                )
            else:
                message = (
                    f"[ms-teams-mcp] Auto-update failed. Run manually:\n"
                    f"  pip install --upgrade ms-teams-mcp"
                )
        else:
            message = None

        # Save cache
        with open(UPDATE_CHECK_CACHE, "w") as f:
            json.dump({
                "last_check": datetime.now(timezone.utc).isoformat(),
                "latest_version": latest_tag,
                "message": message,
            }, f)

        return message
    except Exception:
        return None

# ─────────────────────────────────────────
# MCP Server
# ─────────────────────────────────────────
mcp = FastMCP("Microsoft Teams MCP")

# ═══════════════════════════════════════════
# Teams - Teams/Channels
# ═══════════════════════════════════════════

@mcp.tool()
def list_teams() -> str:
    """List Teams teams that I have joined"""
    data = graph_get("/me/joinedTeams", params={"$select": "id,displayName,description"})
    teams = data.get("value", [])
    if not teams:
        return "No teams found."
    result = []
    for t in teams:
        result.append(f"- {t['displayName']} (desc: {t.get('description','')}) | ID: {t['id']}")
    return "\n".join(result)

@mcp.tool()
def list_channels(team_id: str) -> str:
    """
    List channels of a specific team
    - team_id: Team ID from list_teams
    """
    data = graph_get(f"/teams/{team_id}/channels", params={"$select": "id,displayName,membershipType"})
    channels = data.get("value", [])
    if not channels:
        return "No channels found."
    result = []
    for c in channels:
        result.append(f"- #{c['displayName']} (type: {c.get('membershipType','')}) | ID: {c['id']}")
    return "\n".join(result)

@mcp.tool()
def list_channel_messages(team_id: str, channel_id: str, top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    List messages in a team channel
    - team_id: Team ID
    - channel_id: Channel ID
    - top: Number of messages to retrieve (max 50)
    - skip: Number of messages to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(
            f"/teams/{team_id}/channels/{channel_id}/messages",
            params=params
        )
    messages = data.get("value", [])
    if not messages:
        return "No messages found."
    result = []
    for i, m in enumerate(messages, 1):
        sender_from = m.get("from") or {}
        sender_user = sender_from.get("user") or {}
        sender = sender_user.get("displayName") or "Unknown"
        body = strip_html(m.get("body", {}).get("content", ""))[:200]
        created = m.get("createdDateTime", "")[:19]
        result.append(
            f"{i}. [{created}] {sender}\n"
            f"   {body}\n"
            f"   ID: {m['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_channel_message(team_id: str, channel_id: str, message: str) -> str:
    """
    Send a message to a team channel.
    IMPORTANT: Always show the message content to the user and get explicit confirmation before calling this tool.
    - team_id: Team ID
    - channel_id: Channel ID
    - message: Message content to send
    """
    body = {"body": {"content": message}}
    data = graph_post(f"/teams/{team_id}/channels/{channel_id}/messages", body)
    return f"Message sent (ID: {data.get('id', '')})"

@mcp.tool()
def reply_to_channel_message(team_id: str, channel_id: str, message_id: str, message: str) -> str:
    """
    Reply to a specific message in a team channel.
    IMPORTANT: Always show the reply content to the user and get explicit confirmation before calling this tool.
    - team_id: Team ID
    - channel_id: Channel ID
    - message_id: ID of the message to reply to (from list_channel_messages)
    - message: Reply content
    """
    body = {"body": {"content": message}}
    data = graph_post(f"/teams/{team_id}/channels/{channel_id}/messages/{message_id}/replies", body)
    return f"Reply sent (ID: {data.get('id', '')})"

# ═══════════════════════════════════════════
# Teams - 1:1/Group Chats
# ═══════════════════════════════════════════

@mcp.tool()
def list_chats(top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    List my 1:1/group chats
    - top: Number of chats to retrieve (max 50)
    - skip: Number of chats to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top, "$expand": "members", "$orderby": "lastMessagePreview/createdDateTime desc"}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/chats", params=params)
    chats = data.get("value", [])
    if not chats:
        return "No chats found."
    result = []
    for i, c in enumerate(chats, 1):
        chat_type = c.get("chatType", "")
        topic = c.get("topic") or ""
        members = [m.get("displayName") or "" for m in c.get("members", [])]
        members_str = ", ".join(members[:5])
        preview = c.get("lastMessagePreview", {})
        preview_body = strip_html(preview.get("body", {}).get("content", ""))[:80] if preview else ""
        preview_time = preview.get("createdDateTime", "")[:19] if preview else ""
        label = topic if topic else members_str
        result.append(
            f"{i}. [{chat_type}] {label}\n"
            f"   Latest: [{preview_time}] {preview_body}\n"
            f"   ID: {c['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def list_chat_messages(chat_id: str, top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    List messages in a specific chat
    - chat_id: Chat ID from list_chats
    - top: Number of messages to retrieve (max 50)
    - skip: Number of messages to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(f"/me/chats/{chat_id}/messages", params=params)
    messages = data.get("value", [])
    if not messages:
        return "No messages found."
    result = []
    for i, m in enumerate(messages, 1):
        sender = m.get("from", {})
        sender_name = "System"
        if sender and sender.get("user"):
            sender_name = sender["user"].get("displayName", "Unknown")
        body = strip_html(m.get("body", {}).get("content", ""))[:300]
        created = m.get("createdDateTime", "")[:19]
        msg_type = m.get("messageType", "")
        if msg_type == "systemEventMessage":
            continue
        result.append(
            f"{i}. [{created}] {sender_name}\n"
            f"   {body}\n"
            f"   ID: {m['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_chat_message(chat_id: str, message: str) -> str:
    """
    Send a message to a 1:1/group chat.
    IMPORTANT: Always show the message content to the user and get explicit confirmation before calling this tool.
    - chat_id: Chat ID
    - message: Message content to send
    """
    body = {"body": {"content": message}}
    data = graph_post(f"/me/chats/{chat_id}/messages", body)
    return f"Message sent (ID: {data.get('id', '')})"

@mcp.tool()
def reply_to_chat_message(chat_id: str, message_id: str, message: str) -> str:
    """
    Reply to a specific message in a 1:1/group chat.
    IMPORTANT: Always show the reply content to the user and get explicit confirmation before calling this tool.
    Note: If the Graph API does not support chat message replies, this will send as a regular message in the chat.
    - chat_id: Chat ID
    - message_id: ID of the message to reply to (from list_chat_messages)
    - message: Reply content
    """
    body = {"body": {"content": message}}
    try:
        data = graph_post(f"/me/chats/{chat_id}/messages/{message_id}/replies", body)
        return f"Reply sent (ID: {data.get('id', '')})"
    except Exception as e:
        # Fallback only for API-not-supported errors; re-raise others
        err_str = str(e).lower()
        if "404" in err_str or "not supported" in err_str or "not found" in err_str:
            data = graph_post(f"/me/chats/{chat_id}/messages", body)
            return f"Message sent as regular message (reply API not supported) (ID: {data.get('id', '')})"
        raise

@mcp.tool()
def create_chat(members: str, message: str = "", topic: str = "") -> str:
    """
    Create a new 1:1 or group chat.
    IMPORTANT: Always show the participants and message content to the user and get explicit confirmation before calling this tool.
    - members: Participant email addresses (comma-separated, excluding yourself)
    - message: First message (optional)
    - topic: Group chat topic (optional, recommended for 3+ participants)
    """
    member_list = [m.strip() for m in members.split(",") if m.strip()]
    if not member_list:
        return "Please provide at least one participant email address."
    chat_type = "oneOnOne" if len(member_list) == 1 else "group"
    me = graph_get("/me", params={"$select": "id"})
    body = {
        "chatType": chat_type,
        "members": [
            {
                "@odata.type": "#microsoft.graph.aadUserConversationMember",
                "roles": ["owner"],
                "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{me['id']}')"
            }
        ] + [
            {
                "@odata.type": "#microsoft.graph.aadUserConversationMember",
                "roles": ["owner"],
                "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{email}')"
            }
            for email in member_list
        ]
    }
    if topic and chat_type == "group":
        body["topic"] = topic
    data = graph_post("/chats", body)
    if message:
        graph_post(f"/chats/{data['id']}/messages", {"body": {"content": message}})
    result = f"Chat created\n- ID: {data.get('id', '')}\n- Type: {chat_type}\n- Participants: {', '.join(member_list)}"
    if topic and chat_type == "group":
        result += f"\n- Topic: {topic}"
    if message:
        result += "\n- First message sent"
    return result

# ═══════════════════════════════════════════
# Outlook Mail
# ═══════════════════════════════════════════

@mcp.tool()
def list_emails(folder: str = "inbox", top: int = 10, skip: int = 0, next_link: str = "") -> str:
    """
    List Outlook emails
    - folder: inbox / sentItems / drafts / deleteditems
    - top: Number of emails to retrieve (max 1000)
    - skip: Number of emails to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 1000)
        params = {
            "$top": top,
            "$select": "id,subject,from,receivedDateTime,isRead,bodyPreview",
            "$orderby": "receivedDateTime desc"
        }
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(f"/me/mailFolders/{folder}/messages", params=params)
    emails = data.get("value", [])
    if not emails:
        return "No emails found."
    result = []
    for i, e in enumerate(emails, 1):
        read_mark = "N" if not e.get("isRead") else " "
        sender = e.get("from", {}).get("emailAddress", {})
        result.append(
            f"{i}. [{read_mark}] [{e['receivedDateTime'][:10]}] {e['subject']}\n"
            f"   From: {sender.get('name','')} <{sender.get('address','')}>\n"
            f"   ID: {e['id']}\n"
            f"   Preview: {e.get('bodyPreview','')[:80]}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def read_email(message_id: str) -> str:
    """
    Read full body of a specific email
    - message_id: ID from list_emails
    """
    data = graph_get(f"/me/messages/{message_id}", params={"$select": "subject,from,toRecipients,receivedDateTime,body,isRead"})
    sender = data.get("from", {}).get("emailAddress", {})
    to_list = [r["emailAddress"]["address"] for r in data.get("toRecipients", [])]
    body_text = strip_html(data.get("body", {}).get("content", ""))
    return (
        f"Subject: {data.get('subject')}\n"
        f"From: {sender.get('name')} <{sender.get('address')}>\n"
        f"To: {', '.join(to_list)}\n"
        f"Date: {data.get('receivedDateTime','')[:19]}\n"
        f"{'─'*50}\n"
        f"{body_text[:3000]}"
    )

@mcp.tool()
def search_emails(query: str, top: int = 10, skip: int = 0, next_link: str = "") -> str:
    """
    Search emails
    - query: Search term (subject, body, sender, etc.)
    - top: Number of results (max 1000)
    - skip: Number of results to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 1000)
        params = {
            "$search": f'"{query}"',
            "$top": top,
            "$select": "id,subject,from,receivedDateTime,bodyPreview"
        }
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/messages", params=params)
    emails = data.get("value", [])
    if not emails:
        return f"No results found for '{query}'."
    result = []
    for i, e in enumerate(emails, 1):
        sender = e.get("from", {}).get("emailAddress", {})
        result.append(
            f"{i}. [{e['receivedDateTime'][:10]}] {e['subject']}\n"
            f"   From: {sender.get('name')} <{sender.get('address')}>\n"
            f"   ID: {e['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_email(to: str, subject: str, body: str, cc: str = "", bcc: str = "") -> str:
    """
    Send an email.
    IMPORTANT: Always show the recipients, subject, and body to the user and get explicit confirmation before calling this tool.
    - to: Recipient email (comma-separated for multiple)
    - subject: Email subject
    - body: Email body (text)
    - cc: CC recipients (comma-separated, optional)
    - bcc: BCC recipients (comma-separated, optional)
    """
    message = {
        "subject": subject,
        "body": {"contentType": "Text", "content": body},
        "toRecipients": _parse_recipients(to),
    }
    if cc:
        message["ccRecipients"] = _parse_recipients(cc)
    if bcc:
        message["bccRecipients"] = _parse_recipients(bcc)
    graph_post_action("/me/sendMail", {"message": message})
    return f"Email sent -> {to} (subject: {subject})"

@mcp.tool()
def reply_email(message_id: str, body: str, reply_all: bool = False) -> str:
    """
    Reply to an email.
    IMPORTANT: Always show the reply content to the user and get explicit confirmation before calling this tool.
    - message_id: Original email ID
    - body: Reply content
    - reply_all: True for reply all (default: False)
    """
    action = "replyAll" if reply_all else "reply"
    graph_post_action(f"/me/messages/{message_id}/{action}", {"comment": body})
    action_label = "Reply all" if reply_all else "Reply"
    return f"{action_label} sent (original ID: {message_id})"

@mcp.tool()
def forward_email(message_id: str, to: str, comment: str = "") -> str:
    """
    Forward an email.
    IMPORTANT: Always show the recipients and content to the user and get explicit confirmation before calling this tool.
    - message_id: Original email ID
    - to: Forward recipient email (comma-separated for multiple)
    - comment: Additional comment when forwarding (optional)
    """
    graph_post_action(f"/me/messages/{message_id}/forward", {
        "comment": comment,
        "toRecipients": _parse_recipients(to),
    })
    return f"Email forwarded -> {to} (original ID: {message_id})"

@mcp.tool()
def list_mail_folders() -> str:
    """List available mail folders"""
    data = graph_get("/me/mailFolders", params={"$select": "id,displayName,totalItemCount,unreadItemCount"})
    folders = data.get("value", [])
    result = []
    for f in folders:
        result.append(f"{f['displayName']} (total: {f['totalItemCount']}, unread: {f['unreadItemCount']}) | ID: {f['id']}")
    return "\n".join(result)

# ═══════════════════════════════════════════
# Calendar
# ═══════════════════════════════════════════

@mcp.tool()
def list_calendar_events(start_date: str = "", end_date: str = "", top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    List calendar events within a date range using calendarView.
    - start_date: Start date in YYYY-MM-DD format (default: today)
    - end_date: End date in YYYY-MM-DD format (default: start_date + 7 days)
    - top: Number of events to retrieve (max 50)
    - skip: Number of events to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        if not start_date:
            start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        if not end_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end_date = (start_dt + timedelta(days=7)).strftime("%Y-%m-%d")
        top = min(top, 50)
        params = {
            "startDateTime": f"{start_date}T00:00:00Z",
            "endDateTime": f"{end_date}T23:59:59Z",
            "$top": top,
            "$select": "id,subject,start,end,organizer,location,attendees,isOnlineMeeting,onlineMeetingUrl",
            "$orderby": "start/dateTime",
        }
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/calendarView", params=params)
    events = data.get("value", [])
    if not events:
        return "No calendar events found."
    result = []
    for i, ev in enumerate(events, 1):
        subject = ev.get("subject", "(No subject)")
        start = ev.get("start", {})
        end = ev.get("end", {})
        start_time = start.get("dateTime", "")[:16]
        end_time = end.get("dateTime", "")[:16]
        tz = start.get("timeZone", "")
        organizer = ev.get("organizer", {}).get("emailAddress", {}).get("name", "Unknown")
        location = ev.get("location", {}).get("displayName", "")
        attendees_count = len(ev.get("attendees", []))
        online = "Yes" if ev.get("isOnlineMeeting") else "No"
        line = (
            f"{i}. {subject}\n"
            f"   Time: {start_time} ~ {end_time} ({tz})\n"
            f"   Organizer: {organizer}\n"
        )
        if location:
            line += f"   Location: {location}\n"
        line += (
            f"   Attendees: {attendees_count} | Online meeting: {online}\n"
            f"   ID: {ev['id']}"
        )
        result.append(line)
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def create_calendar_event(subject: str, start: str, end: str, attendees: str = "", location: str = "", body: str = "", is_online: bool = False) -> str:
    """
    Create a new calendar event.
    IMPORTANT: Always show the event details (subject, time, attendees) to the user and get explicit confirmation before calling this tool.
    - subject: Event subject/title
    - start: Start datetime in ISO format (e.g. 2026-03-10T09:00:00)
    - end: End datetime in ISO format (e.g. 2026-03-10T10:00:00)
    - attendees: Attendee email addresses (comma-separated, optional)
    - location: Event location (optional)
    - body: Event description (optional)
    - is_online: Create as Teams online meeting (default: False)
    """
    event = {
        "subject": subject,
        "start": {"dateTime": start, "timeZone": "UTC"},
        "end": {"dateTime": end, "timeZone": "UTC"},
    }
    if attendees:
        event["attendees"] = _parse_attendees(attendees)
    if location:
        event["location"] = {"displayName": location}
    if body:
        event["body"] = {"contentType": "Text", "content": body}
    if is_online:
        event["isOnlineMeeting"] = True
        event["onlineMeetingProvider"] = "teamsForBusiness"
    data = graph_post("/me/events", event)
    result = f"Event created: {subject}\n- ID: {data.get('id', '')}\n- Time: {start} ~ {end}"
    if attendees:
        result += f"\n- Attendees: {attendees}"
    if is_online:
        meeting_url = data.get("onlineMeeting", {}).get("joinUrl", "")
        if meeting_url:
            result += f"\n- Teams meeting URL: {meeting_url}"
    return result

@mcp.tool()
def update_calendar_event(
    event_id: str,
    subject: str = "",
    start: str = "",
    end: str = "",
    attendees: str = "",
    location: str = "",
    body: str = "",
    is_online: str = "",
) -> str:
    """
    Update an existing calendar event.
    IMPORTANT: Always show the updated details to the user and get explicit confirmation before calling this tool.
    - event_id: Event ID (from list_calendar_events)
    - subject: New subject/title (optional, leave empty to keep current)
    - start: New start datetime in ISO format e.g. 2026-03-10T09:00:00 (optional)
    - end: New end datetime in ISO format e.g. 2026-03-10T10:00:00 (optional)
    - attendees: New attendee email addresses, comma-separated (optional, replaces all attendees)
    - location: New location (optional)
    - body: New description (optional)
    - is_online: Set to "true" or "false" to change online meeting setting (optional, leave empty to keep current)
    """
    update = {}
    if subject:
        update["subject"] = subject
    if start:
        update["start"] = {"dateTime": start, "timeZone": "UTC"}
    if end:
        update["end"] = {"dateTime": end, "timeZone": "UTC"}
    if attendees:
        update["attendees"] = _parse_attendees(attendees)
    if location:
        update["location"] = {"displayName": location}
    if body:
        update["body"] = {"contentType": "Text", "content": body}
    if is_online.lower() == "true":
        update["isOnlineMeeting"] = True
        update["onlineMeetingProvider"] = "teamsForBusiness"
    elif is_online.lower() == "false":
        update["isOnlineMeeting"] = False

    if not update:
        return "No fields to update. Provide at least one field to change."

    data = graph_patch(f"/me/events/{event_id}", update)
    updated_subject = data.get("subject", subject or "(unchanged)")
    result = f"Event updated: {updated_subject}\n- ID: {event_id}"
    if start:
        result += f"\n- Start: {data.get('start', {}).get('dateTime', '')[:16]}"
    if end:
        result += f"\n- End: {data.get('end', {}).get('dateTime', '')[:16]}"
    if attendees:
        result += f"\n- Attendees: {attendees}"
    if location:
        result += f"\n- Location: {location}"
    return result

@mcp.tool()
def delete_calendar_event(event_id: str) -> str:
    """
    Delete a calendar event.
    IMPORTANT: Always show the event details to the user and get explicit confirmation before calling this tool.
    - event_id: Event ID (from list_calendar_events)
    """
    graph_delete(f"/me/events/{event_id}")
    return f"Event deleted successfully.\n- ID: {event_id}"

@mcp.tool()
def create_recurring_event(
    subject: str,
    start_time: str,
    end_time: str,
    recurrence_type: str,
    range_start: str,
    range_end: str = "",
    days_of_week: str = "",
    interval: int = 1,
    attendees: str = "",
    location: str = "",
    body: str = "",
    is_online: bool = False,
    occurrences: int = 0,
) -> str:
    """
    Create a recurring calendar event.
    IMPORTANT: Always show the event details to the user and get explicit confirmation before calling this tool.
    - subject: Event subject/title
    - start_time: Event start time in HH:MM format (e.g. 09:00)
    - end_time: Event end time in HH:MM format (e.g. 10:00)
    - recurrence_type: One of "daily", "weekly", "absoluteMonthly", "relativeMonthly", "absoluteYearly"
    - range_start: Recurrence start date in YYYY-MM-DD format (e.g. 2026-03-10)
    - range_end: Recurrence end date in YYYY-MM-DD (optional, use this or occurrences)
    - days_of_week: For weekly recurrence, comma-separated days e.g. "monday,wednesday,friday"
    - interval: Recurrence interval (e.g. 2 = every 2 weeks/days/months, default 1)
    - attendees: Attendee email addresses, comma-separated (optional)
    - location: Event location (optional)
    - body: Event description (optional)
    - is_online: Create as Teams online meeting (default: False)
    - occurrences: Number of occurrences (optional, use this or range_end)
    """
    pattern = {
        "type": recurrence_type,
        "interval": interval,
    }
    if days_of_week and recurrence_type == "weekly":
        pattern["daysOfWeek"] = [d.strip().lower() for d in days_of_week.split(",") if d.strip()]

    rec_range = {"startDate": range_start}
    if range_end:
        rec_range["type"] = "endDate"
        rec_range["endDate"] = range_end
    elif occurrences > 0:
        rec_range["type"] = "numbered"
        rec_range["numberOfOccurrences"] = occurrences
    else:
        rec_range["type"] = "noEnd"

    event = {
        "subject": subject,
        "start": {"dateTime": f"{range_start}T{start_time}:00", "timeZone": "UTC"},
        "end": {"dateTime": f"{range_start}T{end_time}:00", "timeZone": "UTC"},
        "recurrence": {
            "pattern": pattern,
            "range": rec_range,
        },
    }
    if attendees:
        event["attendees"] = _parse_attendees(attendees)
    if location:
        event["location"] = {"displayName": location}
    if body:
        event["body"] = {"contentType": "Text", "content": body}
    if is_online:
        event["isOnlineMeeting"] = True
        event["onlineMeetingProvider"] = "teamsForBusiness"

    data = graph_post("/me/events", event)
    rec_desc = f"{recurrence_type} (every {interval})" if interval > 1 else recurrence_type
    if days_of_week:
        rec_desc += f" on {days_of_week}"
    result = (
        f"Recurring event created: {subject}\n"
        f"- ID: {data.get('id', '')}\n"
        f"- Time: {start_time} ~ {end_time}\n"
        f"- Recurrence: {rec_desc}\n"
        f"- Range: {range_start}"
    )
    if range_end:
        result += f" to {range_end}"
    elif occurrences > 0:
        result += f" ({occurrences} occurrences)"
    else:
        result += " (no end)"
    if attendees:
        result += f"\n- Attendees: {attendees}"
    return result

@mcp.tool()
def create_reminder(
    subject: str,
    remind_at: str,
    body: str = "",
) -> str:
    """
    Create a reminder as a short calendar event with an alert.
    IMPORTANT: Always show the reminder details to the user and get explicit confirmation before calling this tool.
    - subject: Reminder subject/title
    - remind_at: Reminder datetime in ISO format (e.g. 2026-03-10T14:00:00)
    - body: Reminder note/description (optional)
    """
    try:
        end_dt = datetime.fromisoformat(remind_at) + timedelta(minutes=15)
    except ValueError:
        return "Invalid datetime format. Use ISO format e.g. 2026-03-10T14:00:00"
    event = {
        "subject": f"[Reminder] {subject}",
        "start": {"dateTime": remind_at, "timeZone": "UTC"},
        "end": {"dateTime": end_dt.isoformat(), "timeZone": "UTC"},
        "isReminderOn": True,
        "reminderMinutesBeforeStart": 0,
        "showAs": "free",
    }
    if body:
        event["body"] = {"contentType": "Text", "content": body}

    data = graph_post("/me/events", event)
    return (
        f"Reminder created: {subject}\n"
        f"- ID: {data.get('id', '')}\n"
        f"- Time: {remind_at}\n"
        f"- Status: Will alert at the specified time"
    )

# ═══════════════════════════════════════════
# Teams Files (SharePoint-based)
# ═══════════════════════════════════════════

@mcp.tool()
def list_channel_files(team_id: str, channel_id: str, top: int = 50, skip: int = 0, next_link: str = "") -> str:
    """
    List files uploaded to a team channel
    - team_id: Team ID
    - channel_id: Channel ID
    - top: Number of files to retrieve (max 200)
    - skip: Number of files to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        items = graph_get("", url=next_link)
        drive_id = ""
    else:
        data = graph_get(f"/teams/{team_id}/channels/{channel_id}/filesFolder")
        drive_id = data.get("parentReference", {}).get("driveId", "")
        folder_id = data.get("id", "")
        if not drive_id or not folder_id:
            return "Unable to retrieve file folder information."
        top = min(top, 200)
        params = {
            "$select": "id,name,size,lastModifiedDateTime,webUrl,file,folder",
            "$top": top
        }
        if skip > 0:
            params["$skip"] = skip
        items = graph_get(f"/drives/{drive_id}/items/{folder_id}/children", params=params)
    files = items.get("value", [])
    if not files:
        return "No files found."
    # Extract driveId from first file's parentReference when using next_link
    if not drive_id and files:
        drive_id = files[0].get("parentReference", {}).get("driveId", "")
    result = []
    for i, f in enumerate(files, 1):
        ftype = "Folder" if "folder" in f else "File"
        size = f.get("size", 0)
        size_str = f"{size / 1024 / 1024:.1f}MB" if size > 1024 * 1024 else f"{size / 1024:.1f}KB"
        result.append(
            f"{i}. [{ftype}] {f['name']} ({size_str})\n"
            f"   Modified: {f.get('lastModifiedDateTime', '')[:19]}\n"
            f"   ID: {f['id']}\n"
            f"   DriveID: {drive_id}"
        )
    return "\n\n".join(result) + _pagination_footer(items, skip, top)

@mcp.tool()
def read_channel_file(drive_id: str, item_id: str) -> str:
    """
    Read file content from a team channel (text-based files only)
    - drive_id: DriveID from list_channel_files
    - item_id: File ID from list_channel_files
    """
    meta = graph_get(f"/drives/{drive_id}/items/{item_id}", params={"$select": "name,size,file,@microsoft.graph.downloadUrl"})
    name = meta.get("name", "")
    size = meta.get("size", 0)
    download_url = meta.get("@microsoft.graph.downloadUrl", "")

    if size > 5 * 1024 * 1024:
        return f"File too large ({size / 1024 / 1024:.1f}MB). Only files up to 5MB can be read."

    mime = meta.get("file", {}).get("mimeType", "")
    if "image" in mime or "video" in mime or "audio" in mime:
        return f"Binary file ({mime}). Only text files can be read."

    if not download_url:
        return "Unable to retrieve download URL."

    resp = requests.get(download_url)
    _check_response(resp)

    # Handle xlsx files
    if name.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        result_parts = [f"Filename: {name}"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                rows.append(row_str)
                if len(rows) > 200:
                    break
            result_parts.append(f"\n{'─'*50}\n[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        full = "\n".join(result_parts)
        return full[:8000]

    try:
        content = resp.content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            content = resp.content.decode("cp949")
        except UnicodeDecodeError:
            return "Unable to determine file encoding. This may be a binary file."

    return f"Filename: {name}\n{'─'*50}\n{content[:5000]}"

# ═══════════════════════════════════════════
# File Keyword Index
# ═══════════════════════════════════════════

_STOPWORDS = frozenset({
    "the", "and", "for", "that", "this", "with", "from", "are", "was", "were",
    "been", "being", "have", "has", "had", "does", "did", "will", "would",
    "could", "should", "may", "might", "shall", "can", "need", "must",
    "not", "but", "nor", "yet", "also", "just", "than", "then", "too",
    "very", "its", "our", "your", "his", "her", "their", "all", "any",
    "each", "every", "both", "few", "more", "most", "other", "some",
    "such", "only", "own", "same", "into", "over", "after", "before",
    "between", "under", "about", "out", "off", "once", "here", "there",
    "when", "where", "why", "how", "what", "which", "who", "whom",
    "she", "him", "they", "them", "these", "those", "you", "new",
    "use", "used", "using", "one", "two", "get", "set", "com", "www",
    "http", "https", "none", "null", "true", "false", "def", "class",
    "return", "import", "self",
})

def _extract_keywords(text: str, top_n: int = 20) -> list[str]:
    """Extract top-N keywords from text using word frequency"""
    words = re.findall(r'[가-힣]{2,}|[a-zA-Z]{3,}', text)
    filtered = [w.lower() for w in words if w.lower() not in _STOPWORDS]
    if not filtered:
        return []
    return [word for word, _ in Counter(filtered).most_common(top_n)]

@mcp.tool()
def build_file_index() -> str:
    """
    Build a keyword index of all files across all Teams channels.
    Traverses all joined teams and their channels, reads text-based files (up to 5MB),
    extracts keywords, and saves the index to ~/.ms_mcp_file_index.json.
    Use search_file_index to search the built index.
    """
    teams_data = graph_get("/me/joinedTeams", params={"$select": "id,displayName"})
    teams = teams_data.get("value", [])
    if not teams:
        return "No teams found."

    indexed_files = []
    errors = []
    total_scanned = 0

    for team in teams:
        team_id = team["id"]
        team_name = team["displayName"]
        try:
            channels_data = graph_get(f"/teams/{team_id}/channels", params={"$select": "id,displayName"})
        except Exception as e:
            errors.append(f"Failed to list channels for team '{team_name}': {e}")
            continue

        for channel in channels_data.get("value", []):
            channel_id = channel["id"]
            channel_name = channel["displayName"]
            try:
                folder = graph_get(f"/teams/{team_id}/channels/{channel_id}/filesFolder")
                drive_id = folder.get("parentReference", {}).get("driveId", "")
                folder_id = folder.get("id", "")
                if not drive_id or not folder_id:
                    continue
                items = graph_get(
                    f"/drives/{drive_id}/items/{folder_id}/children",
                    params={"$select": "id,name,size,lastModifiedDateTime,file", "$top": 200}
                )
            except Exception as e:
                errors.append(f"Failed to list files for '{team_name}/#{channel_name}': {e}")
                continue

            for item in items.get("value", []):
                if "file" not in item:
                    continue
                total_scanned += 1
                name = item.get("name", "")
                size = item.get("size", 0)
                item_id = item["id"]
                mime = item.get("file", {}).get("mimeType", "")

                # Skip large or binary files
                if size > 5 * 1024 * 1024:
                    continue
                if any(t in mime for t in ("image", "video", "audio")):
                    continue

                try:
                    meta = graph_get(
                        f"/drives/{drive_id}/items/{item_id}",
                        params={"$select": "name,@microsoft.graph.downloadUrl"}
                    )
                    download_url = meta.get("@microsoft.graph.downloadUrl", "")
                    if not download_url:
                        continue

                    resp = requests.get(download_url)
                    if not resp.ok:
                        continue

                    # Parse content
                    if name.endswith(".xlsx"):
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
                        parts = []
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            for row in ws.iter_rows(values_only=True):
                                parts.append(" ".join(str(c) for c in row if c is not None))
                                if len(parts) > 500:
                                    break
                        wb.close()
                        content = " ".join(parts)
                    else:
                        try:
                            content = resp.content.decode("utf-8")
                        except UnicodeDecodeError:
                            try:
                                content = resp.content.decode("cp949")
                            except UnicodeDecodeError:
                                continue

                    keywords = _extract_keywords(content)
                    if not keywords:
                        continue

                    indexed_files.append({
                        "team": team_name,
                        "channel": channel_name,
                        "name": name,
                        "drive_id": drive_id,
                        "item_id": item_id,
                        "size": size,
                        "modified": item.get("lastModifiedDateTime", "")[:19],
                        "keywords": keywords,
                    })
                except Exception as e:
                    errors.append(f"Failed to read '{name}' in '{team_name}/#{channel_name}': {e}")
                    continue

    index = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "files": indexed_files,
    }
    with open(FILE_INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    result = (
        f"File index built successfully.\n"
        f"- Files scanned: {total_scanned}\n"
        f"- Files indexed: {len(indexed_files)}\n"
        f"- Index saved to: {FILE_INDEX_PATH}"
    )
    if errors:
        result += f"\n- Errors: {len(errors)}"
        for err in errors[:5]:
            result += f"\n  * {err}"
    return result

@mcp.tool()
def search_file_index(query: str) -> str:
    """
    Search the file keyword index built by build_file_index.
    - query: Search keywords (e.g. "server IP" or "자산 관리")
    """
    try:
        with open(FILE_INDEX_PATH, "r", encoding="utf-8") as f:
            index = json.load(f)
    except FileNotFoundError:
        return "No file index found. Run build_file_index first to create the index."

    query_words = [w.lower() for w in re.findall(r'[가-힣]{2,}|[a-zA-Z]{3,}', query)]
    if not query_words:
        return "Please provide valid search keywords (2+ Korean chars or 3+ English chars)."

    scored = []
    for entry in index.get("files", []):
        keywords = [k.lower() for k in entry.get("keywords", [])]
        name_words = [w.lower() for w in re.findall(r'[가-힣]{2,}|[a-zA-Z]{3,}', entry.get("name", ""))]
        score = 0
        for qw in query_words:
            # Exact match in keywords
            if qw in keywords:
                score += 2
            # Partial match in keywords
            elif any(qw in k or k in qw for k in keywords):
                score += 1
            # Match in filename
            if qw in name_words:
                score += 3
            elif any(qw in nw or nw in qw for nw in name_words):
                score += 1
        if score > 0:
            scored.append((score, entry))

    scored.sort(key=lambda x: x[0], reverse=True)
    top_results = scored[:10]

    if not top_results:
        return f"No files found matching '{query}'."

    updated = index.get("updated_at", "unknown")
    result = [f"Search results for '{query}' (index updated: {updated[:19]}):\n"]
    for rank, (score, entry) in enumerate(top_results, 1):
        kw_str = ", ".join(entry["keywords"][:10])
        result.append(
            f"{rank}. [{entry['team']}/#{entry['channel']}] {entry['name']}\n"
            f"   Size: {entry['size'] / 1024:.1f}KB | Modified: {entry['modified']}\n"
            f"   Keywords: {kw_str}\n"
            f"   DriveID: {entry['drive_id']} | ItemID: {entry['item_id']}\n"
            f"   Relevance: {score}"
        )
    return "\n\n".join(result)

# ═══════════════════════════════════════════
# User & Summary
# ═══════════════════════════════════════════

@mcp.tool()
def search_users(query: str, top: int = 10, skip: int = 0, next_link: str = "") -> str:
    """
    Search for people in the organization.
    - query: Search term (name, email, etc.)
    - top: Number of results (max 50)
    - skip: Number of results to skip (pagination, default 0)
    - next_link: Next page link (provided by previous result, no manual input needed)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$search": f'"{query}"', "$top": top}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/people", params=params)
    people = data.get("value", [])
    if not people:
        return f"No people found for '{query}'."
    result = []
    for i, p in enumerate(people, 1):
        name = p.get("displayName", "Unknown")
        emails = p.get("scoredEmailAddresses", [])
        email = emails[0].get("address", "") if emails else ""
        title = p.get("jobTitle", "")
        dept = p.get("department", "")
        line = f"{i}. {name}"
        if email:
            line += f" <{email}>"
        if title:
            line += f"\n   Title: {title}"
        if dept:
            line += f"\n   Department: {dept}"
        result.append(line)
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def get_unread_summary() -> str:
    """
    Get a summary of unread emails and recent chat activity.
    Shows inbox unread count and recent chat previews.
    """
    parts = []
    # Unread email count
    try:
        inbox = graph_get("/me/mailFolders/inbox", params={"$select": "unreadItemCount,totalItemCount"})
        unread = inbox.get("unreadItemCount", 0)
        total = inbox.get("totalItemCount", 0)
        parts.append(f"Inbox: {unread} unread / {total} total emails")
    except Exception as e:
        parts.append(f"Inbox: Failed to retrieve ({e})")

    # Recent chats
    try:
        chats = graph_get("/me/chats", params={
            "$top": 10,
            "$expand": "members",
            "$orderby": "lastMessagePreview/createdDateTime desc",
        })
        chat_list = chats.get("value", [])
        if chat_list:
            parts.append(f"\nRecent chats ({len(chat_list)}):")
            for i, c in enumerate(chat_list, 1):
                topic = c.get("topic") or ""
                members = [m.get("displayName") or "" for m in c.get("members", [])]
                members_str = ", ".join(members[:4])
                label = topic if topic else members_str
                preview = c.get("lastMessagePreview", {})
                preview_body = strip_html(preview.get("body", {}).get("content", ""))[:60] if preview else ""
                preview_time = preview.get("createdDateTime", "")[:16] if preview else ""
                parts.append(f"  {i}. {label}\n     [{preview_time}] {preview_body}")
        else:
            parts.append("\nNo recent chats.")
    except Exception as e:
        parts.append(f"\nChats: Failed to retrieve ({e})")

    return "\n".join(parts)

# ═══════════════════════════════════════════
# Auth Management
# ═══════════════════════════════════════════

@mcp.tool()
def check_update() -> str:
    """Check if a newer version of ms-teams-mcp is available on PyPI and auto-update if outdated"""
    try:
        resp = requests.get(
            "https://pypi.org/pypi/ms-teams-mcp/json",
            timeout=5,
        )
        if not resp.ok:
            return f"Failed to check for updates (HTTP {resp.status_code})."

        latest_tag = resp.json().get("info", {}).get("version", "")
        if not latest_tag:
            return f"Current version: {__version__}. No releases found on PyPI."

        latest_ver = _parse_version(latest_tag)
        current_ver = _parse_version(__version__)

        if latest_ver > current_ver:
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--upgrade", "ms-teams-mcp"],
                capture_output=True, text=True, timeout=120,
            )
            if result.returncode == 0:
                return (
                    f"Updated successfully!\n"
                    f"  {__version__} -> {latest_tag}\n"
                    f"  Please restart the server to use the new version."
                )
            return (
                f"Update available but auto-update failed.\n"
                f"  Current: {__version__}\n"
                f"  Latest:  {latest_tag}\n"
                f"  Run manually: pip install --upgrade ms-teams-mcp"
            )
        return f"You are up to date (version {__version__})."
    except Exception as e:
        return f"Failed to check for updates: {e}"

@mcp.tool()
def auth_status() -> str:
    """Check current Microsoft authentication status"""
    app = _get_app()
    accounts = app.get_accounts()
    if not accounts:
        return (
            "Not authenticated - No token found.\n"
            "Call the authenticate tool to authenticate via Device Code Flow."
        )
    account = accounts[0]
    username = account.get("username", "Unknown")
    result = app.acquire_token_silent(SCOPES, account=account)
    if result and "access_token" in result:
        save_cache()
        return f"Authenticated - Account: {username}\nToken is valid."
    else:
        return (
            f"Token expired - Account: {username}\n"
            "Call the authenticate tool to re-authenticate."
        )

_pending_device_flow = None

def _device_code_auth(on_flow_started=None):
    """Device Code Flow common logic. Returns (status, username) on success, raises on failure."""
    # 1. Quick check: only try silent refresh if cached accounts exist
    accounts = _get_app().get_accounts() if os.path.exists(TOKEN_CACHE_FILE) else []
    if accounts:
        result = _get_app().acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            save_cache()
            return ("refreshed", accounts[0]["username"])

    # 2. Start Device Code Flow (skip slow silent refresh, go straight to link)
    pub_app = _get_pub_app()
    flow = pub_app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(f"Failed to start Device Code Flow: {flow.get('error_description', flow)}")

    if on_flow_started:
        on_flow_started(flow)

    # 3. Acquire token
    result = pub_app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        save_cache()
        username = result.get("id_token_claims", {}).get("preferred_username", "Unknown")
        return ("authenticated", username)
    else:
        raise RuntimeError(f"Authentication failed: {result.get('error_description', result)}")

@mcp.tool()
def authenticate() -> str:
    """
    Start authentication via Device Code Flow (Step 1 of 2).
    Returns the login URL and code immediately.
    After visiting the URL and entering the code in your browser, call authenticate_complete to finish.
    """
    global _pending_device_flow

    # Quick check: if already authenticated, return immediately
    try:
        accounts = _get_app().get_accounts() if os.path.exists(TOKEN_CACHE_FILE) else []
        if accounts:
            result = _get_app().acquire_token_silent(SCOPES, account=accounts[0])
            if result and "access_token" in result:
                save_cache()
                return f"Token already valid. Account: {accounts[0]['username']}"
    except Exception:
        pass

    # Start Device Code Flow and return URL/code immediately
    try:
        pub_app = _get_pub_app()
        flow = pub_app.initiate_device_flow(scopes=SCOPES)
        if "user_code" not in flow:
            return f"Failed to start Device Code Flow: {flow.get('error_description', flow)}"

        _pending_device_flow = flow
        return (
            f"Visit this URL to authenticate:\n\n"
            f"  URL:  {flow['verification_uri']}\n"
            f"  Code: {flow['user_code']}\n\n"
            f"After entering the code in your browser, call the authenticate_complete tool to finish."
        )
    except Exception as e:
        return f"Failed to start authentication: {e}"

@mcp.tool()
def authenticate_complete() -> str:
    """
    Complete authentication (Step 2 of 2).
    Call this after you have visited the URL and entered the code from the authenticate tool.
    """
    global _pending_device_flow

    if _pending_device_flow is None:
        return "No pending authentication flow. Call the authenticate tool first."

    flow = _pending_device_flow
    try:
        pub_app = _get_pub_app()
        result = pub_app.acquire_token_by_device_flow(flow)
        _pending_device_flow = None

        if "access_token" in result:
            save_cache()
            username = result.get("id_token_claims", {}).get("preferred_username", "Unknown")
            return f"Authentication complete!\nAccount: {username}\nToken saved to: {TOKEN_CACHE_FILE}"
        else:
            return f"Authentication failed: {result.get('error_description', result)}"
    except Exception as e:
        _pending_device_flow = None
        return f"Authentication failed: {e}"

# ─────────────────────────────────────────
# CLI: Auth subcommand
# ─────────────────────────────────────────

def _parse_auth_args(args):
    """Parse --client-id, --client-secret, --tenant-id for auth subcommand"""
    i = 0
    while i < len(args):
        if args[i] == "--client-id" and i + 1 < len(args):
            os.environ["MS_CLIENT_ID"] = args[i + 1]
            i += 2
        elif args[i] == "--client-secret" and i + 1 < len(args):
            os.environ["MS_CLIENT_SECRET"] = args[i + 1]
            i += 2
        elif args[i] == "--tenant-id" and i + 1 < len(args):
            os.environ["MS_TENANT_ID"] = args[i + 1]
            i += 2
        else:
            i += 1

def cmd_auth():
    """Issue token via Device Code Flow (for headless servers)"""
    def on_flow(flow):
        print("=" * 60, file=sys.stderr, flush=True)
        print("  Microsoft Token Issuance (Device Code Flow)", file=sys.stderr, flush=True)
        print("=" * 60, file=sys.stderr, flush=True)
        print(f"\n  1. Visit on any device: {flow['verification_uri']}", file=sys.stderr, flush=True)
        print(f"  2. Enter code: {flow['user_code']}", file=sys.stderr, flush=True)
        print(f"\n  Expires in: {flow.get('expires_in', 900)} seconds", file=sys.stderr, flush=True)
        print("=" * 60, file=sys.stderr, flush=True)
        print("\nWaiting for authentication...", file=sys.stderr, flush=True)

    print("Connecting to Microsoft identity platform...", file=sys.stderr, flush=True)
    try:
        status, username = _device_code_auth(on_flow_started=on_flow)
    except RuntimeError as e:
        print(str(e))
        sys.exit(1)

    if status == "refreshed":
        print(f"Token refreshed successfully! Account: {username}")
    else:
        print(f"\nToken saved! Account: {username}")
        print(f"Location: {TOKEN_CACHE_FILE}")

def _parse_serve_args(args):
    """Parse --transport, --host, --port for serve subcommand"""
    opts = {"transport": "streamable-http", "host": "127.0.0.1", "port": "7979"}
    i = 0
    while i < len(args):
        if args[i] == "--transport" and i + 1 < len(args):
            opts["transport"] = args[i + 1]
            i += 2
        elif args[i] == "--host" and i + 1 < len(args):
            opts["host"] = args[i + 1]
            i += 2
        elif args[i] == "--port" and i + 1 < len(args):
            opts["port"] = args[i + 1]
            i += 2
        else:
            i += 1
    return opts

def _print_usage():
    print("Usage:")
    print("  ms-teams-mcp                         # Run MCP server (stdio)")
    print("  ms-teams-mcp serve                   # Run MCP server (streamable-http)")
    print("  ms-teams-mcp serve --transport sse   # Run MCP server (SSE)")
    print("  ms-teams-mcp serve --host 0.0.0.0    # Allow external connections")
    print("  ms-teams-mcp serve --port 9000       # Custom port (default: 7979)")
    print("  ms-teams-mcp auth                    # Device Code Flow auth")
    print("  ms-teams-mcp auth \\")
    print("    --client-id <ID> \\")
    print("    --client-secret <SECRET> \\")
    print("    --tenant-id <TENANT>                    # Auth with CLI args")
    print("  ms-teams-mcp --version                # Show version")

def main():
    if len(sys.argv) > 1:
        command = sys.argv[1]
        if command == "auth":
            _parse_auth_args(sys.argv[2:])
            cmd_auth()
        elif command == "serve":
            opts = _parse_serve_args(sys.argv[2:])
            transport = opts["transport"]
            host = opts["host"]
            port = int(opts["port"])
            update_msg = _check_and_auto_update()
            if update_msg and sys.stderr is not None:
                print(update_msg, file=sys.stderr, flush=True)
            print(f"Starting MCP server ({transport}) on {host}:{port}...")
            mcp.run(transport=transport, host=host, port=port)
        elif command in ("--version", "-v", "version"):
            print(f"ms-teams-mcp {__version__}")
        else:
            _print_usage()
            sys.exit(1)
    else:
        update_msg = _check_and_auto_update()
        if update_msg and sys.stderr is not None:
            print(update_msg, file=sys.stderr, flush=True)
        print("Starting Microsoft Teams MCP server...")
        mcp.run()

if __name__ == "__main__":
    main()
